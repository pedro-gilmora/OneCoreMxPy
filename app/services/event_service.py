"""
Event Log Service for historical tracking of system events.
"""
import json
from datetime import datetime
from typing import Optional, List, BinaryIO
from io import BytesIO
from functools import lru_cache

from sqlalchemy.orm import Session
from sqlalchemy import and_, or_

from app.models.models import EventLog, EventType, User
from app.schemas.schemas import (
    EventLogCreate,
    EventLogResponse,
    EventLogFilter,
    EventTypeEnum,
)


class EventService:
    """Service class for event logging and history management."""
    
    def create_event(
        self,
        db: Session,
        event_type: EventTypeEnum,
        description: str,
        user_id: Optional[int] = None,
        document_id: Optional[int] = None,
        metadata: Optional[dict] = None
    ) -> EventLog:
        """
        Create a new event log entry.
        
        Args:
            db: Database session
            event_type: Type of event
            description: Description of the event
            user_id: Optional user ID
            document_id: Optional document ID
            metadata: Optional additional metadata
            
        Returns:
            Created EventLog instance
        """
        event = EventLog(
            event_type=event_type.value,
            description=description,
            user_id=user_id,
            document_id=document_id,
            metadata_json=json.dumps(metadata) if metadata else None,
        )
        
        db.add(event)
        db.commit()
        db.refresh(event)
        
        return event
    
    def get_event(self, db: Session, event_id: int) -> Optional[EventLog]:
        """Get a single event by ID."""
        return db.query(EventLog).filter(EventLog.id == event_id).first()
    
    def get_events(
        self,
        db: Session,
        filters: Optional[EventLogFilter] = None,
        page: int = 1,
        page_size: int = 20
    ) -> tuple[List[EventLog], int]:
        """
        Get events with optional filtering and pagination.
        
        Args:
            db: Database session
            filters: Optional filters to apply
            page: Page number (1-indexed)
            page_size: Number of items per page
            
        Returns:
            Tuple of (list of events, total count)
        """
        query = db.query(EventLog)
        
        # Apply filters
        if filters:
            conditions = []
            
            if filters.event_type:
                conditions.append(EventLog.event_type == filters.event_type.value)
            
            if filters.description_search:
                conditions.append(
                    EventLog.description.ilike(f"%{filters.description_search}%")
                )
            
            if filters.date_from:
                conditions.append(EventLog.created_at >= filters.date_from)
            
            if filters.date_to:
                conditions.append(EventLog.created_at <= filters.date_to)
            
            if filters.user_id:
                conditions.append(EventLog.user_id == filters.user_id)
            
            if conditions:
                query = query.filter(and_(*conditions))
        
        # Get total count
        total = query.count()
        
        # Apply pagination and ordering
        offset = (page - 1) * page_size
        events = query.order_by(EventLog.created_at.desc()).offset(offset).limit(page_size).all()
        
        return events, total
    
    def event_to_response(self, event: EventLog, db: Session) -> EventLogResponse:
        """Convert EventLog model to response schema."""
        username = None
        if event.user_id:
            user = db.query(User).filter(User.id == event.user_id).first()
            if user:
                username = user.username
        
        metadata = None
        if event.metadata_json:
            try:
                metadata = json.loads(event.metadata_json)
            except json.JSONDecodeError:
                metadata = None
        
        return EventLogResponse(
            id=event.id,
            event_type=event.event_type,
            description=event.description,
            document_id=event.document_id,
            user_id=event.user_id,
            username=username,
            metadata=metadata,
            created_at=event.created_at
        )
    
    def export_to_excel(
        self,
        db: Session,
        filters: Optional[EventLogFilter] = None
    ) -> BytesIO:
        """
        Export filtered events to Excel format.
        
        Args:
            db: Database session
            filters: Optional filters to apply
            
        Returns:
            BytesIO buffer containing the Excel file
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Get all events (no pagination for export)
        events, _ = self.get_events(db, filters, page=1, page_size=100000)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Histórico de Eventos"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ["ID", "Tipo", "Descripción", "Usuario", "Documento ID", "Fecha y Hora"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Event type labels
        event_type_labels = {
            "subida_documento": "Subida de Documento",
            "analisis_ia": "Análisis IA",
            "interaccion_usuario": "Interacción Usuario",
            "sistema": "Sistema"
        }
        
        # Data rows
        for row, event in enumerate(events, 2):
            username = None
            if event.user_id:
                user = db.query(User).filter(User.id == event.user_id).first()
                if user:
                    username = user.username
            
            ws.cell(row=row, column=1, value=event.id).border = thin_border
            ws.cell(row=row, column=2, value=event_type_labels.get(event.event_type, event.event_type)).border = thin_border
            ws.cell(row=row, column=3, value=event.description).border = thin_border
            ws.cell(row=row, column=4, value=username or "-").border = thin_border
            ws.cell(row=row, column=5, value=event.document_id or "-").border = thin_border
            ws.cell(row=row, column=6, value=event.created_at.strftime("%Y-%m-%d %H:%M:%S")).border = thin_border
        
        # Adjust column widths
        column_widths = [10, 25, 60, 20, 15, 22]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    # Convenience methods for common event types
    def log_document_upload(
        self,
        db: Session,
        document_id: int,
        filename: str,
        user_id: int
    ) -> EventLog:
        """Log a document upload event."""
        return self.create_event(
            db=db,
            event_type=EventTypeEnum.DOCUMENT_UPLOAD,
            description=f"Documento '{filename}' subido exitosamente",
            user_id=user_id,
            document_id=document_id,
            metadata={"filename": filename}
        )
    
    def log_ai_analysis(
        self,
        db: Session,
        document_id: int,
        document_type: str,
        user_id: Optional[int] = None,
        success: bool = True,
        error: Optional[str] = None
    ) -> EventLog:
        """Log an AI analysis event."""
        if success:
            description = f"Análisis IA completado. Documento clasificado como: {document_type}"
        else:
            description = f"Error en análisis IA: {error}"
        
        return self.create_event(
            db=db,
            event_type=EventTypeEnum.AI_ANALYSIS,
            description=description,
            user_id=user_id,
            document_id=document_id,
            metadata={
                "document_type": document_type,
                "success": success,
                "error": error
            }
        )
    
    def log_user_interaction(
        self,
        db: Session,
        action: str,
        user_id: int,
        document_id: Optional[int] = None,
        details: Optional[dict] = None
    ) -> EventLog:
        """Log a user interaction event."""
        return self.create_event(
            db=db,
            event_type=EventTypeEnum.USER_INTERACTION,
            description=f"Interacción de usuario: {action}",
            user_id=user_id,
            document_id=document_id,
            metadata=details
        )
    
    def log_system_event(
        self,
        db: Session,
        description: str,
        metadata: Optional[dict] = None
    ) -> EventLog:
        """Log a system event."""
        return self.create_event(
            db=db,
            event_type=EventTypeEnum.SYSTEM,
            description=description,
            metadata=metadata
        )


@lru_cache()
def get_event_service() -> EventService:
    """Get cached event service instance."""
    return EventService()
